import os
import datetime as dt
from pathlib import Path
from tkinter import Tk, Button, Label, filedialog, messagebox
from openpyxl import load_workbook

replaceDict = {
    # replace with your own requirements
}

SUPPORTED_EXTS = [".xlsx", ".xlsm"]


def timestamp() -> str:
    return dt.datetime.now().strftime("%Y%m%d-%H%M%S")


def process_xlsx_xlsm(path: Path, rep_dict: dict) -> dict:
    """Process .xlsx or .xlsm ."""
    keep_vba = path.suffix.lower() == ".xlsm"
    wb = load_workbook(filename=str(path), keep_vba=keep_vba, data_only=False)
    counter = {k: 0 for k in rep_dict.keys()}

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val in rep_dict.keys():
                    cell.value = rep_dict[val]
                    counter[val] += 1

    wb.save(str(path))
    return counter


def process_workbook(file_path: str, rep_dict: dict) -> dict:
    """Detect extension and route to appropriate processor. Returns replaceCountDict."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    ext = path.suffix.lower()
    if ext in SUPPORTED_EXTS:
        return process_xlsx_xlsm(path, rep_dict)
    else:
        raise ValueError("Unsupported file type. Use .xlsx, .xlsm")


def process_path(path_like, rep_dict: dict, recursive: bool = True):
    """
    Process a single Excel file or every supported Excel file in a folder.
    `process_workbook(file_path: str, rep_dict: dict)` and `SUPPORTED_EXTS`.

    Returns a summary dict:
    {
      "mode": "file" | "folder",
      "files_processed": int,
      "files_succeeded": int,
      "files_failed": int,
      "total_counts": {key: int, ...},
      "per_file_counts": { "file.xlsx": {key: int, ...}, ... },
      "failures": [("file.xlsb", "ErrorMessage"), ...]
    }
    """
    p = Path(path_like)
    if not p.exists():
        raise FileNotFoundError(f"Path not found: {p}")

    def is_supported_excel(f: Path) -> bool:
        # Skip Excel temp/lock files
        if f.name.startswith("~$"):
            return False
        return f.suffix.lower() in SUPPORTED_EXTS

    total_counts = {k: 0 for k in rep_dict.keys()}
    per_file_counts = {}
    failures = []

    if p.is_file():
        counts = process_workbook(str(p), rep_dict)
        per_file_counts[p.name] = counts
        for k, v in counts.items():
            total_counts[k] += v
        return {
            "mode": "file",
            "files_processed": 1,
            "files_succeeded": 1,
            "files_failed": 0,
            "total_counts": total_counts,
            "per_file_counts": per_file_counts,
            "failures": failures,
        }

    # Folder mode
    files = p.rglob("*") if recursive else p.glob("*")
    targets = [f for f in files if f.is_file() and is_supported_excel(f)]

    succeeded = 0
    for f in sorted(targets):
        try:
            counts = process_workbook(str(f), rep_dict)
            per_file_counts[str(f)] = counts
            for k, v in counts.items():
                total_counts[k] += v
            succeeded += 1
        except Exception as e:
            failures.append((str(f), f"{type(e).__name__}: {e}"))

    return {
        "mode": "folder",
        "files_processed": len(targets),
        "files_succeeded": succeeded,
        "files_failed": len(failures),
        "total_counts": total_counts,
        "per_file_counts": per_file_counts,
        "failures": failures,
    }


# ------------------ Tkinter GUI ------------------


class App:  # Tested on Windows Only
    def __init__(self, root):
        root.title("Excel Replace Automation")
        root.geometry("700x240")
        root.resizable(False, False)

        self.label = Label(root, text="Select an Excel file (.xlsx, .xlsm) to process:")
        self.label.pack(pady=10)

        self.button = Button(root, text="Choose File and Run", command=self.run)
        self.button.pack(pady=5)

        self.status = Label(root, text="", fg="gray")
        self.status.pack(pady=5)

        self.folder_btn = Button(
            root,
            text="Choose Folder and Run (recursive)",
            command=self.choose_folder_and_run,
        )
        self.folder_btn.pack(pady=2)

    def choose_folder_and_run(self):
        folder = filedialog.askdirectory(title="Select folder containing Excel files")
        if not folder:
            return
        self.status.config(text="Running (folder)...")
        self.button.config(state="disabled")
        self.folder_btn.config(state="disabled")
        root.update_idletasks()

        try:
            summary = process_path(folder, replaceDict, recursive=True)
            total = sum(summary["total_counts"].values())
            lines = [
                "Batch complete.",
                f"Mode: {summary['mode']}",
                f"Files processed: {summary['files_processed']}",
                f"Succeeded: {summary['files_succeeded']} | Failed: {summary['files_failed']}",
                f"Total cells changed: {total}",
            ]
            nz = {k: v for k, v in summary["total_counts"].items() if v}
            if nz:
                lines.append("Total per-key counts:")
                width = max(len(str(k)) for k in nz)
                for k in sorted(nz):
                    lines.append(f"  {str(k).ljust(width)} : {nz[k]}")
            if summary["failures"]:
                lines.append("\nFailures:")
                for fpath, err in summary["failures"][:10]:                           
                    lines.append(f"  {fpath} -> {err}")

            self.status.config(text="Done.")
            messagebox.showinfo("Done", "\n".join(lines))
        except Exception as e:
            self.status.config(text="Error.")
            messagebox.showerror("Error", f"{type(e).__name__}: {e}")
        finally:
            self.button.config(state="normal")
            self.folder_btn.config(state="normal")

    def run(self):
        filetypes = [
            ("Excel files", "*.xlsx *.xlsm"),
            ("All files", "*.*"),
        ]
        filename = filedialog.askopenfilename(
            title="Select Excel file", filetypes=filetypes
        )
        if not filename:
            return

        self.status.config(text="Running... please wait.")
        self.button.config(state="disabled")
        root.update_idletasks()

        try:
            counts = process_workbook(filename, replaceDict)
            total_replacements = sum(counts.values())
            nonzero = {k: v for k, v in counts.items() if v}
            lines = [
                f"Replacements complete.",
                f"File: {Path(filename).name}",
                f"Total replacements: {total_replacements}",
            ]
            if nonzero:
                lines.append("Per-key counts:")
                width = max(len(k) for k in nonzero.keys())
                for k, v in sorted(nonzero.items()):
                    lines.append(f"  {k.ljust(width)} : {v}")
            else:
                lines.append("No keys were found.")

            self.status.config(text="Done.")
            messagebox.showinfo("Automation Complete", "\n".join(lines))

        except Exception as e:
            self.status.config(text="Error.")
            messagebox.showerror("Error", f"{type(e).__name__}: {e}")

        finally:
            self.button.config(state="normal")


if __name__ == "__main__":
    root = Tk()
    App(root)
    root.mainloop()
